import openpyxl

# Criar uma planilha(book)
book = openpyxl.Workbook()

# Como visualizar páginas existentes
print(book.sheetnames)

# Como criar uma página
book.create_sheet('Frutas')

# Como selecionar uma página
frutas_page = book['Frutas']

# Como adicionar dados a página

frutas_page.append(['Frutas', 'Quantidade', 'Preço'])
frutas_page.append(['Banana', '5', 'R$3,90'])
frutas_page.append(['Manga', '2', 'R$15,90'])
frutas_page.append(['Melancia', '10', 'R$30,90'])
frutas_page.append(['Morango', '15', 'R$50,50'])

# Salvar a planilha
book.save('Planilha de Compras.xlsx')